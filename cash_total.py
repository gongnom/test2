import openpyxl
from datetime import datetime

def process_list(input_list):
    """
    리스트를 받아서 지정된 형식의 리스트를 반환합니다.
    
    Args:
    input_list (list): [발급수단번호1, 총거래금액, 메모]
    
    Returns:
    list: [0, 발급수단번호2, 총거래금액, 공급가액, 부가가치세, 메모]
    """
    발급수단번호1, 총거래금액, 메모 = input_list
    
    # 발급수단번호2 생성
    발급수단번호2 = f"010-0000-{발급수단번호1[-4:]}"
    
    # 부가가치세 계산 (소수점 이하 버림)
    부가가치세 = 총거래금액 // 11
    
    # 공급가액 계산
    공급가액 = 총거래금액 - 부가가치세
    
    return [0, 발급수단번호2, 총거래금액, 공급가액, 부가가치세, 메모]




def save_to_excel(data):
    """
    중복된 리스트 데이터를 엑셀 파일에 저장합니다. 7행부터 입력합니다.
    
    Args:
    data (list): 변환된 리스트들의 리스트
    """
    # 오늘 날짜를 "년월일" 형식으로 생성
    today_date = datetime.now().strftime("%y%m%d")
    file_name = f"cash_{today_date}.xlsx"
    
    # 워크북 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # 헤더 작성 (6행에 작성)
    headers = ["Index", "발급수단번호2", "총거래금액", "공급가액", "부가가치세", "메모"]
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=6, column=col_num, value=header)  # 헤더를 6행에 작성
    
    # 7행부터 데이터 작성
    start_row = 7
    for idx, row in enumerate(data, start=0):  # 데이터 작성
        for col_num, value in enumerate([0] + row[1:], start=1):  # 1열에 항상 '0' 추가
            sheet.cell(row=start_row + idx, column=col_num, value=value)
    
    # 엑셀 파일 저장
    workbook.save(file_name)
    print(f"데이터가 {file_name} 파일에 저장되었습니다.")



# 예제 데이터
input_data = [
    ['1234', 110000, '메모1'],
    ['5678', 220000, '메모2'],
    ['9101', 330000, '메모3']
]

# 변환된 데이터 저장
processed_data = [process_list(item) for item in input_data]

# 엑셀로 저장
save_to_excel(processed_data)

